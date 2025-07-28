import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import Counter

def make_unique_column_names(column_names):
    """
    重複するカラム名に連番を付けてユニークにする
    
    Parameters:
    column_names: カラム名のリスト
    
    Returns:
    unique_names: ユニークなカラム名のリスト
    """
    seen = {}
    unique_names = []
    
    for name in column_names:
        if name in seen:
            seen[name] += 1
            unique_name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
            unique_name = name
        
        unique_names.append(unique_name)
    
    return unique_names

def extract_visible_columns_with_merged_headers_unique(file_path, sheet_name=None, col_range=None, 
                                                     header_row1=2, header_row2=3, separator='_'):
    """
    グループ化されていない列を抽出し、セル結合された複数行のヘッダーを結合してカラム名とする
    重複するカラム名には自動で連番を付与
    
    Parameters:
    file_path: Excelファイルのパス
    sheet_name: シート名（Noneの場合はアクティブシート）
    col_range: 列の範囲 (例: 'A:J' または (1, 10))
    header_row1: 1番目のヘッダー行（1ベース）
    header_row2: 2番目のヘッダー行（1ベース）
    separator: ヘッダー結合時の区切り文字
    """
    
    # Excelファイルを読み込み
    workbook = load_workbook(file_path)
    if sheet_name:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active
    
    # 列範囲を決定
    if col_range:
        if isinstance(col_range, str):
            start_col = ord(col_range.split(':')[0]) - ord('A') + 1
            end_col = ord(col_range.split(':')[1]) - ord('A') + 1
        else:
            start_col, end_col = col_range
    else:
        start_col = 1
        end_col = worksheet.max_column
    
    # 表示されている列を特定
    visible_columns = []
    for col_num in range(start_col, end_col + 1):
        col_letter = get_column_letter(col_num)
        column_dimension = worksheet.column_dimensions.get(col_letter)
        
        if column_dimension is None or not column_dimension.hidden:
            visible_columns.append(col_num)
    
    # 結合セルの情報を取得する関数
    def get_merged_cell_value(row, col):
        """指定したセルが結合セルの場合、その値を取得"""
        cell = worksheet.cell(row=row, column=col)
        
        # 結合セルかどうかをチェック
        for merged_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 結合セルの左上のセルの値を取得
                top_left_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                return top_left_cell.value
        
        # 結合セルでない場合は通常の値を返す
        return cell.value
    
    # 各表示列のヘッダー値を取得
    header1_values = []
    header2_values = []
    
    for col_num in visible_columns:
        # 1行目のヘッダー値を取得
        header1_val = get_merged_cell_value(header_row1, col_num)
        header1_values.append(str(header1_val) if header1_val is not None else '')
        
        # 2行目のヘッダー値を取得
        header2_val = get_merged_cell_value(header_row2, col_num)
        header2_values.append(str(header2_val) if header2_val is not None else '')
    
    workbook.close()
    
    # カラム名を結合して作成
    combined_headers = []
    for i, (h1, h2) in enumerate(zip(header1_values, header2_values)):
        if h1 and h2:
            combined_header = f"{h1}{separator}{h2}"
        elif h1:
            combined_header = h1
        elif h2:
            combined_header = h2
        else:
            combined_header = f"Column_{i + 1}"
        
        combined_headers.append(combined_header)
    
    # カラム名をユニークにする
    unique_headers = make_unique_column_names(combined_headers)
    
    # pandasでExcelファイルを読み込み（データ行から開始）
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, skiprows=max(header_row1, header_row2))
    
    # 表示列のみを選択
    visible_col_indices = [col - 1 for col in visible_columns]  # 0ベースに変換
    visible_df = df.iloc[:, visible_col_indices]
    
    # ユニークなカラム名を設定
    visible_df.columns = unique_headers
    
    return visible_df, visible_columns, header1_values, header2_values, combined_headers, unique_headers

# より詳細な処理を含む高機能版
def extract_with_advanced_merged_headers_unique(file_path, sheet_name=None, col_range=None, 
                                              header_rows=[2, 3], separator='_', data_start_row=4,
                                              empty_header_prefix='Column'):
    """
    複数の結合ヘッダー行に対応した高機能版（重複カラム名処理付き）
    
    Parameters:
    header_rows: ヘッダー行のリスト
    data_start_row: データ開始行
    empty_header_prefix: 空のヘッダーの場合のプレフィックス
    """
    
    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    
    # 列範囲を決定
    if col_range:
        if isinstance(col_range, str):
            start_col = ord(col_range.split(':')[0]) - ord('A') + 1
            end_col = ord(col_range.split(':')[1]) - ord('A') + 1
        else:
            start_col, end_col = col_range
    else:
        start_col = 1
        end_col = worksheet.max_column
    
    # 表示されている列を特定
    visible_columns = []
    for col_num in range(start_col, end_col + 1):
        col_letter = get_column_letter(col_num)
        column_dimension = worksheet.column_dimensions.get(col_letter)
        
        if column_dimension is None or not column_dimension.hidden:
            visible_columns.append(col_num)
    
    # 結合セルのマッピングを作成
    def create_merged_cell_mapping():
        """結合セルの情報をマッピングとして作成"""
        merged_mapping = {}
        for merged_range in worksheet.merged_cells.ranges:
            top_left_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            value = top_left_cell.value
            
            # 結合範囲内のすべてのセルに同じ値をマッピング
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_mapping[(row, col)] = value
        
        return merged_mapping
    
    merged_mapping = create_merged_cell_mapping()
    
    # 各ヘッダー行の値を取得
    all_header_values = []
    for row_num in header_rows:
        row_values = []
        for col_num in visible_columns:
            # 結合セルマッピングから値を取得、なければ通常のセル値
            if (row_num, col_num) in merged_mapping:
                value = merged_mapping[(row_num, col_num)]
            else:
                value = worksheet.cell(row=row_num, column=col_num).value
            
            row_values.append(str(value) if value is not None else '')
        all_header_values.append(row_values)
    
    workbook.close()
    
    # カラム名を結合
    combined_headers = []
    for i in range(len(visible_columns)):
        header_parts = [row_values[i] for row_values in all_header_values if row_values[i]]
        
        if header_parts:
            combined_header = separator.join(header_parts)
        else:
            combined_header = f"{empty_header_prefix}_{i + 1}"
        
        combined_headers.append(combined_header)
    
    # カラム名をユニークにする
    unique_headers = make_unique_column_names(combined_headers)
    
    # データを読み込み
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, skiprows=data_start_row - 1)
    
    # 表示列のみを選択
    visible_col_indices = [col - 1 for col in visible_columns]
    visible_df = df.iloc[:, visible_col_indices]
    visible_df.columns = unique_headers
    
    return visible_df, visible_columns, all_header_values, combined_headers, unique_headers

# デバッグと分析用の関数
def analyze_column_duplicates(column_names):
    """
    カラム名の重複状況を分析
    """
    counter = Counter(column_names)
    duplicates = {name: count for name, count in counter.items() if count > 1}
    
    print("=== カラム名重複分析 ===")
    print(f"総カラム数: {len(column_names)}")
    print(f"ユニークなカラム名数: {len(counter)}")
    print(f"重複があるカラム名数: {len(duplicates)}")
    
    if duplicates:
        print("\n重複しているカラム名:")
        for name, count in duplicates.items():
            print(f"  '{name}': {count}回")
    
    return duplicates

# 使用例とデモ
def demo_usage_with_duplicates():
    """使用例とデバッグ情報の表示（重複処理付き）"""
    
    file_path = 'data.xlsx'
    sheet_name = 'Sheet1'
    
    try:
        # 基本版の使用
        result = extract_visible_columns_with_merged_headers_unique(
            file_path, sheet_name, col_range='A:M', separator='_'
        )
        
        df, visible_cols, header1, header2, original_headers, unique_headers = result
        
        print("=== 抽出結果（重複処理付き） ===")
        print(f"表示されている列数: {len(visible_cols)}")
        print(f"表示列のインデックス: {visible_cols}")
        
        print(f"\n2行目ヘッダー: {header1}")
        print(f"3行目ヘッダー: {header2}")
        
        print(f"\n元のカラム名（重複あり）:")
        for i, col_name in enumerate(original_headers):
            print(f"  {i+1}. {col_name}")
        
        # 重複分析
        analyze_column_duplicates(original_headers)
        
        print(f"\n最終的なユニークカラム名:")
        for i, col_name in enumerate(unique_headers):
            print(f"  {i+1}. {col_name}")
        
        print(f"\nデータフレームの形状: {df.shape}")
        print(f"\n最初の5行:")
        print(df.head())
        
        return df
        
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        return None

# より簡潔な使用例
def process_excel_with_unique_headers(file_path, sheet_name=None, **kwargs):
    """
    Excelファイルを処理してユニークなヘッダー付きDataFrameを返す簡易版
    """
    df, *details = extract_visible_columns_with_merged_headers_unique(
        file_path, sheet_name, **kwargs
    )
    
    print(f"処理完了: {df.shape[0]}行 × {df.shape[1]}列")
    print(f"カラム名: {list(df.columns)}")
    
    return df

# 実行例
if __name__ == "__main__":
    # 使用例1: 基本的な使用
    df1 = process_excel_with_unique_headers(
        'data.xlsx', 
        sheet_name='Sheet1',
        col_range='A:M',
        separator='_'
    )
    
    # 使用例2: 詳細な情報付き
    # result_df = demo_usage_with_duplicates()
